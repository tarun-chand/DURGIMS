from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import ProductCategoryMaster,ProductDetails,ProductCompanyMaster,ProductModelMaster,HealthStatusDetails
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl



@login_required(login_url="user_login")
def newPuchaseRedirect(request):    
    list_all_pc = ProductCategoryMaster.objects.all()    
    return render(request,"purchase/newPurchase.html",{'list_all_pc':list_all_pc})

@login_required(login_url="user_login")
def getSerialNumber(request):        
    path = "/media/ubuntu/myZone/durgcourt/durginventory/IMS/myworkshop/"+request.GET.get('srnumber')
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    mrow = sheet_obj.max_row    
    sr_list = []
    for i in range(2, mrow + 1):
        sr_list.append(sheet_obj.cell(row = i, column = 1).value)
    print("srnumber------",sr_list)        
    return render(request,"ajaxpage/purchaseAJX.html",{'flag':'SerialNo','sr_list':sr_list})       


