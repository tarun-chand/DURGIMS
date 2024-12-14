from openpyxl import load_workbook
import psycopg2 
conn = psycopg2.connect(host="localhost",database='durg_ims_pdev',user='postgres',password='123')
cursor =conn.cursor()
path = 'myworkshop/health.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active
max_row = len(sheet_obj['A'])
print(max_row)
max_col = 5
l = []

# for r in range(1,max_row+1):
#     col=[]
#     for c in range(1,max_col+1):
#         cell_obj = sheet_obj.cell(row=r,column=c)
#         col.insert(c,cell_obj.value)
#     l.append(tuple(col))
# print(l)
# try:
#     q = "insert into inventory_transactiondetails(trans_issue_date,issue_received_status,sdm_id) values(%s,%s,%s)"
#     print(q)
#     cursor.executemany(q,l)
#     conn.commit()
#     conn.close()
#     print("DONE")
# except Exception as e:
#     print("ERROR -- ",e)



for r in range(1,max_row+1):
    col=[]
    for c in range(1,max_col+1):
        cell_obj = sheet_obj.cell(row=r,column=c)
        col.insert(c,cell_obj.value)
    l.append(tuple(col))
print(l)
try:
    q = "insert into inventory_healthstatusdetails(product_healthy,health_remarks,product_status,status_date,product_id) values(%s,%s,%s,%s,%s)"
    print(q)
    cursor.executemany(q,l)
    conn.commit()
    conn.close()
    print("DONE")
except Exception as e:
    print("ERROR -- ",e)


# for r in range(1,max_row+1):
#     col=[]
#     for c in range(1,max_col+1):
#         cell_obj = sheet_obj.cell(row=r,column=c)
#         col.insert(c,cell_obj.value)
#     l.append(tuple(col))
# print(l)
# try:
#     q = "insert into inventory_productdetails(product_serialno,entry_date,initial_quantity,current_quantity,cartridge_toner,remarks,product_new_or_open,product_mod_id,propur_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
#     print(q)
#     # cursor.executemany(q,l)
#     # conn.commit()
#     # conn.close()
#     print("DONE")
# except Exception as e:
#     print("ERROR -- ",e)
# truncate table store_transactiondetails RESTART IDENTITY CASCADE ;

