from itertools import product
from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import ProductCategoryMaster,ProductDetails,ProductCompanyMaster,ProductModelMaster,HealthStatusDetails,PurchaseDetails,ProductPuchaseDetails
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
from django.conf import settings
from DURG_IMS.settings import MEDIA_URL,BASE_DIR,MEDIA_ROOT


@login_required(login_url="user_login")
def newPuchaseRedirect(request):    
    list_all_pc = ProductCategoryMaster.objects.all()    
    return render(request,"purchase/newPurchase.html",{'list_all_pc':list_all_pc})

@login_required(login_url="user_login")
def getSerialNumber(request):      
    # path =  BASE_DIR+"/myworkshop/"+request.GET.get('srnumber')
    path =  str(BASE_DIR)+"/myworkshop/"+request.GET.get('srnumber')    
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    mrow = sheet_obj.max_row    
    sr_list = []
    for i in range(1, mrow + 1):
        sr_list.append(sheet_obj.cell(row = i, column = 1).value)
    print("srnumber------",sr_list)        
    return render(request,"ajaxpage/purchaseAJX.html",{'flag':'SerialNo','sr_list':sr_list})       


@login_required(login_url="user_login")
def newPuchaseSubmit(request):    
  # **********------ STORING MEMO/BILL PDF IN FOLDER -----**********
  billmemo=request.FILES['billmemofile']  
  final = MEDIA_ROOT+"/BillMemo/"+billmemo.name
  with open(final, "wb+") as destination:
    for chunk in billmemo.chunks():
      destination.write(chunk) 

  # **********------ STORING DATA IN PURCHASE TABLE -----**********

  purchase_date = request.POST.get('purchase_date')
  pur = PurchaseDetails()   
  pur.purchase_date =  purchase_date
  pur.billmemono = request.POST.get('bill_memo_no') 
  pur.purchased_by = request.POST.get('purchased_by') 
  pur.purchased_for = request.POST.get('purchased_for') 
  pur.purchase_remarks = request.POST.get('purchase_remarks')   
  pur.billmemodoc = billmemo.name
#   pur.save()

  prodet = request.POST.getlist('prodet')
  proRem = request.POST.getlist('proRem')  
  boxdetails = request.FILES.getlist('boxdetails') 
  producttype = request.POST.getlist('product_type')
  print("producttype-------",producttype)
  no_of_item = request.POST.getlist('no_of_item')
  last_pur_id = (PurchaseDetails.objects.last()).purchase_id
  bic = 0
  for ind in range(len(prodet)):    
    pro_pur = ProductPuchaseDetails()  
    if producttype[ind] == 'NONCONS':
        # **********------ STORING BOX DETAILS JPEG IN FOLDER -----**********
        final = MEDIA_ROOT+"/productBoxDetails/"+boxdetails[bic].name
        with open(final, "wb+") as destination:
            for chunk in boxdetails[bic].chunks():
                destination.write(chunk)  
        pro_pur.box_detail = boxdetails[bic].name  
        bic = bic+1
    else:
        pro_pur.box_detail = ""

  # **********------ STORING DATA IN PRODUCT PURCHASE TABLE -----**********    
      
    pro_pur.propur_remarks= proRem[ind]
    pro_pur.purchase = PurchaseDetails.objects.get(purchase_id=last_pur_id) 
    pro_pur.no_of_item = no_of_item[ind]
    pro_pur.product = ProductModelMaster.objects.get(product_mod_id = prodet[ind])
    # pro_pur.save()
   
    # **********------ STORING DATA IN PRODUCT TABLE -----**********
        
    last_propur_id = (ProductPuchaseDetails.objects.last()).propur_id
    # LOGIC FOR CONSUMABLE AND NON CONSUMABLE PRODUCT
    
    if producttype[ind] == 'CONS':    
        pro_det = ProductDetails()
        pro_det.entry_date = purchase_date
        pro_det.product_serialno = ''
        pro_det.initial_quantity = no_of_item[ind]
        pro_det.current_quantity = no_of_item[ind]
        pro_det.cartridge_toner = ''
        pro_det.product_new_or_open='NEW'
        pro_det.product_mod = ProductModelMaster.objects.get(product_mod_id = prodet[ind])
        pro_det.propur = ProductPuchaseDetails.objects.get(propur_id=last_propur_id) 
        # pro_det.save()
    elif producttype[ind] == 'NONCONS':  
        print("ELSE producttype-------",producttype[ind],ind)      
        a = "proserno"+str(ind+1)
        proserno = request.POST.getlist(a)    
        print(proserno)
        for si,srno in enumerate(proserno):
            pro_det = ProductDetails()
            pro_det.product_serialno = srno
            pro_det.entry_date = purchase_date
            pro_det.initial_quantity = 1
            pro_det.current_quantity = 1

            if request.POST.get('tonarName') == "NOT_A_PRINTER":
                pro_det.cartridge_toner = ''
            else:
                pro_det.cartridge_toner = request.POST.get('tonarName')

            pro_det.product_new_or_open='NEW'
            pro_det.product_mod = ProductModelMaster.objects.get(product_mod_id = prodet[ind])
            pro_det.propur = ProductPuchaseDetails.objects.get(propur_id=last_propur_id) 
            # pro_det.save()
  return render(request,"purchase/newPurchase.html")
