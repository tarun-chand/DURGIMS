from openpyxl import load_workbook

path = 'datadumper.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active
max_row = len(sheet_obj['A'])
l=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54]
print(len(l))

col1=[]
col2=[]
for r in range(0,len(l)):
  sheet_obj.cell(row=r+1, column=1, value=r+1)
  sheet_obj.cell(row=r+1, column=2, value=l[r])
  sheet_obj.cell(row=r+1, column=3, value=37)
  print(r+1," ",l[r])
  wb_obj.save(path)

# for r in range(0,len(l)):
#   col1.append(l[r])
#   col2.append(37)
#   for c in range(1,9):
#       # print("DATA = ",l[r],"c = ",c)
#       col1.append(l[r])
#       col2.append(c)
       
# for r in range(1,296):  
#   sheet_obj.cell(row=r+1, column=2, value=col1[r])
#   sheet_obj.cell(row=r+1, column=3, value=col2[r])
#   print(col1[r]," ",col2[r])
#   wb_obj.save(path)
